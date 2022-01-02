from datetime import date, datetime
import locale
from pprint import pprint

from openpyxl import load_workbook

from ..data import LegalEntity, LegalEntityPreviousAddress, \
    LegalEntityPreviousName, LegalEntityMediaMention, LegalEntitySource, \
    Owner, OtherRelationship, Person, PersonSource


def load_excel(excel_path):
    wb = load_workbook(filename=excel_path)

    legal_entities = load_legal_entities(wb)
    people = load_people(wb)
    owners = load_owners(wb, legal_entities, people)
    other_relationships = load_other_relationships(wb, legal_entities, people)
    legal_entities_previous_names = load_legal_entities_previous_names(wb, legal_entities)
    legal_entities_previous_addresses = load_legal_entities_previous_addresses(wb, legal_entities)
    legal_entities_sources = load_legal_entities_sources(wb, legal_entities)
    legal_entities_media_mentions = load_legal_entities_media_mentions(wb, legal_entities)
    people_sources = load_people_sources(wb, people)

    return {
        'legal_entities': legal_entities,
        'people': people,
        'owners': owners,
        'other_relationships': other_relationships,
        'legal_entities_previous_names': legal_entities_previous_names,
        'legal_entities_previous_addresses': legal_entities_previous_addresses,
        'legal_entities_sources': legal_entities_sources,
        'legal_entities_media_mentions': legal_entities_media_mentions,
        'people_sources': people_sources
    }

def load_legal_entities(wb):
    ws = wb['1. Legal entities']

    cols_map = {
        'Database identifier': 'database_identifier',
        'Legal entity type': 'legal_entity_type',
        'Name': 'name',
        'Country': 'country',
        'Identification number': 'identification_number',
        'Address': 'address',
        'Foundation date': 'foundation_date',
        'Dissolution date': 'dissolution_date',
        'Other notes': 'other_notes'
    }

    cols_by_number = get_col_sheet_names_by_number(ws, cols_map)

    legal_entities = []

    data_rows = ws.iter_rows(min_row=5, max_row=ws.max_row)
    for row in data_rows:
        legal_entity_data = {}

        for cell in row:
            if cell.column in cols_by_number and cell.value is not None:
                col_sheet_name = cols_by_number[cell.column]

                value = cell.value

                if isinstance(value, str):
                    value = value.strip()

                if col_sheet_name == 'Foundation date' or col_sheet_name == 'Dissolution date':
                    value = parse_date(value)

                legal_entity_data[cols_map[col_sheet_name]] = value

        if legal_entity_data:
            legal_entities.append(LegalEntity(**legal_entity_data))

    legal_entities = sorted(legal_entities, key=lambda le: locale.strxfrm(le.name))

    return legal_entities

def load_people(wb):
    ws = wb['2. People']

    cols_map = {
        'Database identifier': 'database_identifier',
        'Full name': 'full_name',
        'Nationality': 'nationality',
        'Date of birth': 'date_of_birth',
        'Residence country': 'residence_country',
        'Residence full address': 'residence_address',
        'Residence only city': 'residence_city',
        'Other notes': 'other_notes'
    }
    cols_by_number = get_col_sheet_names_by_number(ws, cols_map)

    people = []

    data_rows = ws.iter_rows(min_row=5, max_row=ws.max_row)
    for row in data_rows:
        person_data = {}

        for cell in row:
            if cell.column in cols_by_number and cell.value is not None:
                people_col = cols_by_number[cell.column]

                value = cell.value

                if isinstance(value, str):
                    value = value.strip()

                if people_col == 'Date of birth':
                    value = parse_date(value)

                person_data[cols_map[people_col]] = value

        if person_data:
            people.append(Person(**person_data))

    people = sorted(people, key=lambda p: tuple(reversed(list(map(lambda x: locale.strxfrm(x), p.full_name.split(' '))))))

    return people


def load_owners(wb, legal_entities, people):
    sheet_name = '1.1. Legal entities owners'
    ws = wb[sheet_name]

    cols_map = {
        'Owned legal entity reference': 'owned_legal_entity',
        'Owner legal entity or person reference': 'owner_legal_entity_or_person',
        'Owner type': 'owner_type',
        'Owner name': 'owner_name',
        'Owner country': 'owner_country',
        'Owner address': 'owner_address',
        'Owner legal entity identification number': 'owner_legal_entity_identification_number',
        'Owner person date of birth': 'owner_person_date_of_birth',
        'Owned percentage': 'owned_percentage',
        'Owned since date': 'owned_since_date',
        'Owned until date': 'owned_until_date',
        'Ownership details': 'ownership_details'
    }
    cols_by_number = get_col_sheet_names_by_number(ws, cols_map)

    owners = []

    data_rows = ws.iter_rows(min_row=5, max_row=ws.max_row)
    for row in data_rows:
        owner_data = {}

        for cell in row:
            if cell.column in cols_by_number and cell.value is not None:
                col_sheet_name = cols_by_number[cell.column]

                value = cell.value

                if isinstance(value, str):
                    value = value.strip()

                if col_sheet_name in ['Owned legal entity reference']:
                    # TODO: what if empty?

                    found_legal_entity = next((le for le in legal_entities if le.database_identifier == value), None)

                    if found_legal_entity is None:
                        raise Exception(f'Ownership relationship in row {cell.row} of sheet "{sheet_name}" is in column "{col_sheet_name}" referencing legal entity "{value}", but no legal entity with such Database identifier exists.')

                    value = found_legal_entity

                if col_sheet_name in ['Owner legal entity or person reference'] and value:
                    found_legal_entity = next((le for le in legal_entities if le.database_identifier == value), None)
                    found_person = next((p for p in people if p.database_identifier == value), None)

                    if found_legal_entity is None and found_person is None:
                        raise Exception(f'Ownership relationship in row {cell.row} of sheet "{sheet_name}" is in column "{col_sheet_name}" referencing legal entity or person "{value}", but no legal entity or person with such Database identifier exists.')
    
                    value = found_legal_entity if found_legal_entity else found_person

                if col_sheet_name in ['Owned since date', 'Owned until date', 'Owner person date of birth']:
                    value = parse_date(value)

                owner_data[cols_map[col_sheet_name]] = value

        if owner_data:
            owners.append(Owner(**owner_data))

    return owners


def load_other_relationships(wb, legal_entities, people):
    # Excel cannot have sheet names longer than 31 chars
    sheet_name = '1.2. Legal entities other relationships'[0:31]

    ws = wb[sheet_name]

    cols_map = {
        'Legal entity reference': 'legal_entity',
        'Related legal entity or person reference': 'related_legal_entity_or_person',
        'Related type': 'related_type',
        'Related name': 'related_name',
        'Related country': 'related_country',
        'Related address': 'related_address',
        'Related legal entity identification number': 'related_legal_entity_identification_number',
        'Related person date of birth': 'related_person_date_of_birth',
        'Related since date': 'related_since_date',
        'Related until date': 'related_until_date',
        'Relationship details': 'relationship_details'
    }
    cols_by_number = get_col_sheet_names_by_number(ws, cols_map)

    other_relationships = []

    data_rows = ws.iter_rows(min_row=5, max_row=ws.max_row)
    for row in data_rows:
        relationship_data = {}

        for cell in row:
            if cell.column in cols_by_number and cell.value is not None:
                col_sheet_name = cols_by_number[cell.column]

                value = cell.value

                if isinstance(value, str):
                    value = value.strip()

                if col_sheet_name in ['Legal entity reference']:
                    # TODO: what if empty?

                    found_legal_entity = next((le for le in legal_entities if le.database_identifier == value), None)

                    if found_legal_entity is None:
                        raise Exception(f'Other relationship in row {cell.row} of sheet "{sheet_name}" is in column "{col_sheet_name}" referencing legal entity "{value}", but no legal entity with such Database identifier exists.')

                    value = found_legal_entity

                if col_sheet_name in ['Related legal entity or person reference'] and value:
                    found_legal_entity = next((le for le in legal_entities if le.database_identifier == value), None)
                    found_person = next((p for p in people if p.database_identifier == value), None)

                    if found_legal_entity is None and found_person is None:
                        raise Exception(f'Other relationship in row {cell.row} of sheet "{sheet_name}" is in column "{col_sheet_name}" referencing legal entity or person "{value}", but no legal entity or person with such Database identifier exists.')
    
                    value = found_legal_entity if found_legal_entity else found_person

                if col_sheet_name in ['Related since date', 'Related until date', 'Related person date of birth']:
                    value = parse_date(value)

                relationship_data[cols_map[col_sheet_name]] = value

        if relationship_data:
            other_relationships.append(OtherRelationship(**relationship_data))

    return other_relationships


def load_legal_entities_previous_names(wb, legal_entities):
    # Excel cannot have sheet names longer than 31 chars
    sheet_name = '1.4. Legal entities previous names'[0:31]

    ws = wb[sheet_name]

    cols_map = {
        'Legal entity reference': 'legal_entity',
        'Previous name': 'previous_name',
        'Named since date': 'named_since_date',
        'Named until date': 'named_until_date'
    }
    cols_by_number = get_col_sheet_names_by_number(ws, cols_map)

    previous_names = []

    data_rows = ws.iter_rows(min_row=5, max_row=ws.max_row)
    for row in data_rows:
        previous_name_data = {}

        for cell in row:
            if cell.column in cols_by_number and cell.value is not None:
                col_sheet_name = cols_by_number[cell.column]

                value = cell.value

                if isinstance(value, str):
                    value = value.strip()

                if col_sheet_name in ['Legal entity reference']:
                    # TODO: what if empty?

                    found_legal_entity = next((le for le in legal_entities if le.database_identifier == value), None)

                    if found_legal_entity is None:
                        raise Exception(f'Legal entity previous name in row {cell.row} of sheet "{sheet_name}" is in column "{col_sheet_name}" referencing legal entity "{value}", but no legal entity with such Database identifier exists.')

                    value = found_legal_entity

                if col_sheet_name in ['Named since date', 'Named until date']:
                    value = parse_date(value)

                previous_name_data[cols_map[col_sheet_name]] = value

        if previous_name_data:
            previous_names.append(LegalEntityPreviousName(**previous_name_data))

    return previous_names


def load_legal_entities_previous_addresses(wb, legal_entities):
    # Excel cannot have sheet names longer than 31 chars
    sheet_name = '1.5. Legal entities previous addresses'[0:31]

    ws = wb[sheet_name]

    cols_map = {
        'Legal entity reference': 'legal_entity',
        'Previous address': 'previous_address',
        'Address since date': 'address_since_date',
        'Address until date': 'address_until_date'
    }
    cols_by_number = get_col_sheet_names_by_number(ws, cols_map)

    previous_addresses = []

    data_rows = ws.iter_rows(min_row=5, max_row=ws.max_row)
    for row in data_rows:
        previous_address_data = {}

        for cell in row:
            if cell.column in cols_by_number and cell.value is not None:
                col_sheet_name = cols_by_number[cell.column]

                value = cell.value

                if isinstance(value, str):
                    value = value.strip()

                if col_sheet_name in ['Legal entity reference']:
                    # TODO: what if empty?

                    found_legal_entity = next((le for le in legal_entities if le.database_identifier == value), None)

                    if found_legal_entity is None:
                        raise Exception(f'Legal entity previous address in row {cell.row} of sheet "{sheet_name}" is in column "{col_sheet_name}" referencing legal entity "{value}", but no legal entity with such Database identifier exists.')

                    value = found_legal_entity

                if col_sheet_name in ['Address since date', 'Address until date']:
                    value = parse_date(value)

                previous_address_data[cols_map[col_sheet_name]] = value

        if previous_address_data:
            previous_addresses.append(LegalEntityPreviousAddress(**previous_address_data))

    return previous_addresses


def load_legal_entities_sources(wb, legal_entities):
    sheet_name = '1.3. Legal entities sources'

    ws = wb[sheet_name]

    cols_map = {
        'Legal entity reference': 'legal_entity',
        'Source summary': 'summary',
        'Information gained from source': 'information_gained_from_source',
        'Source last checked date': 'last_checked_date',
        'Source URL': 'url'
    }
    cols_by_number = get_col_sheet_names_by_number(ws, cols_map)

    sources = []

    data_rows = ws.iter_rows(min_row=5, max_row=ws.max_row)
    for row in data_rows:
        source_data = {}

        for cell in row:
            if cell.column in cols_by_number and cell.value is not None:
                col_sheet_name = cols_by_number[cell.column]

                value = cell.value

                if isinstance(value, str):
                    value = value.strip()

                if col_sheet_name in ['Legal entity reference']:
                    # TODO: what if empty?

                    found_legal_entity = next((le for le in legal_entities if le.database_identifier == value), None)

                    if found_legal_entity is None:
                        raise Exception(f'Legal entity source in row {cell.row} of sheet "{sheet_name}" is in column "{col_sheet_name}" referencing legal entity "{value}", but no legal entity with such Database identifier exists.')

                    value = found_legal_entity

                if col_sheet_name in ['Source last checked date']:
                    value = parse_date(value)

                source_data[cols_map[col_sheet_name]] = value

        if source_data:
            sources.append(LegalEntitySource(**source_data))

    return sources


def load_legal_entities_media_mentions(wb, legal_entities):
    # Excel cannot have sheet names longer than 31 chars
    sheet_name = '1.6. Legal entities media mentions'[0:31]

    ws = wb[sheet_name]

    cols_map = {
        'Legal entity reference': 'legal_entity',
        'Summary of the media mention': 'summary',
        'Media last checked date': 'last_checked_date',
        'Media mention url': 'url'
    }
    cols_by_number = get_col_sheet_names_by_number(ws, cols_map)

    media_mentions = []

    data_rows = ws.iter_rows(min_row=5, max_row=ws.max_row)
    for row in data_rows:
        media_mention_data = {}

        for cell in row:
            if cell.column in cols_by_number and cell.value is not None:
                col_sheet_name = cols_by_number[cell.column]

                value = cell.value

                if isinstance(value, str):
                    value = value.strip()

                if col_sheet_name in ['Legal entity reference']:
                    # TODO: what if empty?

                    found_legal_entity = next((le for le in legal_entities if le.database_identifier == value), None)

                    if found_legal_entity is None:
                        raise Exception(f'Legal entity media mention in row {cell.row} of sheet "{sheet_name}" is in column "{col_sheet_name}" referencing legal entity "{value}", but no legal entity with such Database identifier exists.')

                    value = found_legal_entity

                if col_sheet_name in ['Media last checked date']:
                    value = parse_date(value)

                media_mention_data[cols_map[col_sheet_name]] = value

        if media_mention_data:
            media_mentions.append(LegalEntityMediaMention(**media_mention_data))

    return media_mentions


def load_people_sources(wb, people):
    # Excel cannot have sheet names longer than 31 chars
    sheet_name = '2.1. People sources'[0:31]

    ws = wb[sheet_name]

    cols_map = {
        'Person reference': 'person',
        'Source summary': 'summary',
        'Information gained from source': 'information_gained_from_source',
        'Source last checked date': 'last_checked_date',
        'Source URL': 'url'
    }
    cols_by_number = get_col_sheet_names_by_number(ws, cols_map)

    sources = []

    data_rows = ws.iter_rows(min_row=5, max_row=ws.max_row)
    for row in data_rows:
        source_data = {}

        for cell in row:
            if cell.column in cols_by_number and cell.value is not None:
                col_sheet_name = cols_by_number[cell.column]

                value = cell.value

                if isinstance(value, str):
                    value = value.strip()

                if col_sheet_name in ['Person reference']:
                    # TODO: what if empty?

                    found_person = next((p for p in people if p.database_identifier == value), None)

                    if found_person is None:
                        raise Exception(f'Person source in row {cell.row} of sheet "{sheet_name}" is in column "{col_sheet_name}" referencing person "{value}", but no person with such Database identifier exists.')

                    value = found_person

                if col_sheet_name in ['Media last checked date']:
                    value = parse_date(value)

                source_data[cols_map[col_sheet_name]] = value

        if source_data:
            sources.append(PersonSource(**source_data))

    return sources


def get_col_sheet_names_by_number(ws, cols_map):
    col_sheet_names_by_number = {}
    header_row = ws.iter_rows(min_row=4, max_row=4)
    
    for cell in list(header_row)[0]:
        for col_sheet_name in cols_map.keys():
            if cell.value == col_sheet_name:
                col_sheet_names_by_number[cell.column] = col_sheet_name
    
    return col_sheet_names_by_number


def parse_date(date_str):
    if date_str == "":
        return None

    if isinstance(date_str, datetime):
        return date_str.date()

    datetime_obj = None

    try:
        datetime_obj = datetime.strptime(date_str, '%Y-%m-%d')
    except ValueError:
        pass

    try:
        datetime_obj = datetime.strptime(date_str, '%d.%m.%Y')
    except ValueError:
        pass

    return datetime_obj.date() if datetime_obj else None
