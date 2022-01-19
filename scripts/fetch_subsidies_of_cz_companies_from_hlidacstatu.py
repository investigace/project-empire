#!/usr/bin/env python3

import argparse
from os import getenv
from os.path import join, dirname, abspath
from pprint import pprint
import sys

import openpyxl
from prompt_toolkit.shortcuts import confirm, prompt

import empire

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Fetch subsidies of CZ companies in Empire database from Hlidacstatu.cz')
    parser.add_argument(
        'database_excel',
        metavar="DATABASE_EXCEL",
        help='Path to the Excel spreadsheet file with Empire database'
    )

    if len(sys.argv) == 1:
        parser.print_help(sys.stderr)
        sys.exit(1)

    args = parser.parse_args()

    hlidacstatu_auth_token = prompt(
        message=f'Hlidacstatu AUTH token: ',
        is_password=True
    )

    out_excel_path = abspath(join(dirname(__file__), 'subsidies_of_cz_companies_from_hlidacstatu.xlsx'))

    # Load data

    empire_data = empire.load_excel(args.database_excel)

    print(f"Loaded Empire database: {len(empire_data['legal_entities'])} legal entities, {len(empire_data['people'])} people, {len(empire_data['subsidies'])} subsidies")

    # Check connection to Hlidacstatu

    hlidacstatu = empire.Hlidacstatu(hlidacstatu_auth_token)

    print('Connected to Hlidacstatu with the provided AUTH token')

    # Filter out CZ legal entites with identifier

    cz_legal_entites = list(filter(lambda c: c.country == 'CZ' and c.identification_number is not None, empire_data['legal_entities']))

    print(f"Will fetch subsidies for {len(cz_legal_entites)} legal entities from CZ country which have identification number filled")

    # Check if continue

    answer = confirm("Do you want to continue?")
    if answer != True:
        print('Ok, exiting')
        exit(0)

    # Fetch and save to Excel file

    fetched_empire_data = hlidacstatu.fetch_subsidies(cz_legal_entites)

    # pprint(len(fetched_empire_data['subsidies']))
    # pprint(len(fetched_empire_data['subsidies_payments']))
    # pprint(len(fetched_empire_data['subsidies_sources']))
    # exit(1)

    wb = openpyxl.Workbook()
    
    subsidies_sheet = wb.active
    subsidies_sheet.title = '3. Subsidies'
    subsidies_sheet.append([
        'Database identifier',
        'Recieving legal entity reference',
        'Year',
        'Project name',
        'Project code',
        'Programme name',
        'Programme code',
        'Notes'
    ])
    for subsidy in fetched_empire_data['subsidies']:
        subsidies_sheet.append([
            subsidy.database_identifier,
            subsidy.receiving_legal_entity.database_identifier,
            subsidy.year,
            subsidy.project_name,
            subsidy.project_code,
            subsidy.programme_name,
            subsidy.programme_code,
            subsidy.notes
        ])

    subsidies_payments_sheet = wb.create_sheet(title="3.1. Subsidies payments", index=1)
    subsidies_payments_sheet.append([
        'Subsidy reference',
        'Provider',
        'Year',
        'Original currency',
        'Amount in original currency',
        'Amount in EUR',
        'Notes'
    ])
    for subsidy_payment in fetched_empire_data['subsidies_payments']:
        subsidies_payments_sheet.append([
            subsidy_payment.subsidy.database_identifier,
            subsidy_payment.provider,
            subsidy_payment.year,
            subsidy_payment.original_currency,
            subsidy_payment.amount_in_original_currency,
            subsidy_payment.amount_in_eur,
            subsidy_payment.notes
        ])

    subsidies_sources_sheet = wb.create_sheet(title="3.2. Subsidies sources", index=2)
    subsidies_sources_sheet.append([
        'Subsidy reference',
        'Source summary',
        'Information gained from source',
        'Source last checked date',
        'Source URL'
    ])
    for subsidy_source in fetched_empire_data['subsidies_sources']:
        subsidies_sources_sheet.append([
            subsidy_source.subsidy.database_identifier,
            subsidy_source.summary,
            subsidy_source.information_gained_from_source,
            subsidy_source.last_checked_date,
            subsidy_source.url
        ])

    wb.save(out_excel_path)

    print(f'Subsidies fetched and saved to Excel file {out_excel_path}')
