#!/usr/bin/env python3

import argparse
import json
from os import getenv
from os.path import join, dirname, abspath
from pprint import pprint
import sys
from tkinter import E

import openpyxl
from prompt_toolkit.shortcuts import confirm, prompt

import empire

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Fetch relationships of CZ companies in Empire database from ARES')
    parser.add_argument(
        'database_excel',
        metavar="DATABASE_EXCEL",
        help='Path to the Excel spreadsheet file with Empire database'
    )
    parser.add_argument(
        '-o',
        '--only-companies-without-relationships',
        help='Fetch for only CZ companies which do not have currently any owner or other relationships',
        action='store_true'
    )
    parser.add_argument(
        '-c',
        '--cache-ares-xmls',
        help='Cache XMLs downloaded from ARES API in system temporary dir so they are downloaded once when you run the script multiple times',
        action='store_true'
    )
    parser.add_argument(
        '-y',
        '--yes',
        help='Do not ask whether to continue and just continue with fetching',
        action='store_true'
    )

    if len(sys.argv) == 1:
        parser.print_help(sys.stderr)
        sys.exit(1)

    args = parser.parse_args()

    # Load data

    print('Loading Empire database...')

    empire_data = empire.load_excel(args.database_excel)

    print(f"Loaded Empire database: {len(empire_data['legal_entities'])} legal entities, {len(empire_data['people'])} people, {len(empire_data['subsidies'])} subsidies")

    # Filter out CZ legal entites with identifier and optionally by relationships

    cz_legal_entites = list(filter(lambda c: c.country == 'CZ' and c.identification_number is not None, empire_data['legal_entities']))

    if args.only_companies_without_relationships:
        cz_legal_entites_without_relationships = []
        
        for legal_entity in cz_legal_entites:
            owners = list(filter(lambda o: o.owned_legal_entity == legal_entity, empire_data['owners']))
            other_relationships = list(filter(lambda o: o.legal_entity == legal_entity, empire_data['other_relationships']))

            if len(owners) == 0 and len(other_relationships) == 0:
                cz_legal_entites_without_relationships.append(legal_entity)

        cz_legal_entites = cz_legal_entites_without_relationships

        print(f"Will fetch relationships for {len(cz_legal_entites)} legal entities from CZ country which have identification number filled and which does not have any relationships now")
    else:
        print(f"Will fetch relationships for {len(cz_legal_entites)} legal entities from CZ country which have identification number filled")

    # Check if continue

    if not args.yes:
        answer = confirm("Do you want to continue?")
        if answer != True:
            print('Ok, exiting')
            exit(0)

    # Do the fetching

    ares = empire.Ares()
    fetched_empire_data = ares.fetch_relationships(cz_legal_entites, cache_ares_xmls=args.cache_ares_xmls)

    # Link to people

    fetched_empire_data = ares.link_fetched_relationships(empire_data, fetched_empire_data)

    # Write to Excel

    out_excel_path = abspath(join(dirname(__file__), 'relationships_of_cz_companies_from_ares.xlsx'))

    wb = openpyxl.Workbook()
    
    owners_sheet = wb.active
    owners_sheet.title = '1.1. Legal entities owners'
    owners_sheet.append([
        'Owned legal entity reference',
        'Owner legal entity or person reference',
        'Owner type',
        'Owner name',
        'Owner country',
        'Owner address',
        'Owner legal entity identification number',
        'Owner person date of birth',
        'Owned percentage',
        'Owned since date',
        'Owned until date',
        'Ownership details'
    ])
    for owner in fetched_empire_data['legal_entities_owners']:
        owners_sheet.append([
            owner.owned_legal_entity.database_identifier,
            owner.owner_legal_entity_or_person.database_identifier if owner.owner_legal_entity_or_person else None,
            owner.owner_type,
            owner.owner_name,
            owner.owner_country,
            owner.owner_address,
            owner.owner_legal_entity_identification_number,
            owner.owner_person_date_of_birth,
            owner.owned_percentage,
            owner.owned_since_date,
            owner.owned_until_date,
            owner.ownership_details
        ])

    # Excel sheet name can be max 31 chars long
    other_relationships_sheet = wb.create_sheet(title="1.2. Legal entities other relat", index=1)
    other_relationships_sheet.append([
        'Legal entity reference',
        'Related legal entity or person reference',
        'Related type',
        'Related name',
        'Related country',
        'Related address',
        'Related legal entity identification number',
        'Related person date of birth',
        'Related since date',
        'Related until date',
        'Relationship details'
    ])
    for other_relationship in fetched_empire_data['legal_entities_other_relationships']:
        other_relationships_sheet.append([
            other_relationship.legal_entity.database_identifier,
            other_relationship.related_legal_entity_or_person.database_identifier if other_relationship.related_legal_entity_or_person else None,
            other_relationship.related_type,
            other_relationship.related_name,
            other_relationship.related_country,
            other_relationship.related_address,
            other_relationship.related_legal_entity_identification_number,
            other_relationship.related_person_date_of_birth,
            other_relationship.related_since_date,
            other_relationship.related_until_date,
            other_relationship.relationship_details
        ])

    people_sheet = wb.create_sheet(title="2. People", index=2)
    people_sheet.append([
        'Database identifier',
        'Full name',
        'Nationality',
        'Date of birth',
        'Residence country',
        'Residence full address',
        'Residence only city',
        'Other notes'
    ])
    for person in fetched_empire_data['people']:
        people_sheet.append([
            person.database_identifier,
            person.full_name,
            person.nationality,
            person.date_of_birth,
            person.residence_country,
            person.residence_address,
            person.residence_city,
            person.other_notes
        ])

    wb.save(out_excel_path)

    print(f'Relationships fetched and saved to Excel file {out_excel_path}')
